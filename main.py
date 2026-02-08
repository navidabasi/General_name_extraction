"""
Names Generation - Main Entry Point

A modular name extraction system for processing Ventrata and Monday booking data.

Features:
- Platform-specific extractors (GYG Standard, GYG MDA, Non-GYG)
- Comprehensive error validation
- DOB/age-based unit type assignment
- Youth validation for EU countries
- Duplicate detection
- Forbidden content checking

Usage:
    python main.py --ventrata path/to/ventrata.xlsx [--monday path/to/monday.xlsx]
"""

import logging
import sys
import os
from pathlib import Path

import pandas as pd

from data_loader import load_ventrata, load_monday, load_update_file, merge_data
from processor import NameExtractionProcessor


def get_next_available_filename(base_filename):
    """
    Get the next available filename by appending a number if file exists.
    
    Args:
        base_filename: Base filename (e.g., 'names_output.xlsx')
        
    Returns:
        str: Available filename (e.g., 'names_output_1.xlsx')
        
    Example:
        If names_output.xlsx exists -> returns names_output_1.xlsx
        If names_output_1.xlsx exists -> returns names_output_2.xlsx
    """
    if not os.path.exists(base_filename):
        return base_filename
    
    # Split filename and extension
    base_path = Path(base_filename)
    name_without_ext = base_path.stem
    extension = base_path.suffix
    directory = base_path.parent if base_path.parent.name else Path('.')
    
    # Find next available number
    counter = 1
    while True:
        new_filename = directory / f"{name_without_ext}_{counter}{extension}"
        if not new_filename.exists():
            return str(new_filename)
        counter += 1


# Set up logging
# Write logs to Documents folder to avoid permission issues on macOS
documents_path = os.path.join(os.path.expanduser('~'), 'Documents')
log_file_path = os.path.join(documents_path, 'namesgen.log')
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    handlers=[
        logging.FileHandler(log_file_path),
        logging.StreamHandler()
    ]
)

logger = logging.getLogger(__name__)


def save_results_to_excel(results_df, output_file, update_row_colors=None):
    """
    Save results to Excel with formatting.
    
    Features:
    - Merged cells for Order Reference and Error columns (same booking)
    - Preserved row colors from update file (when update_row_colors provided)
    - Yellow highlighting for rows with errors
    - Faded yellow highlighting for Youth converted to Adult (non-EU)
    - Alternating row colors (gray/white); skipped for rows with update-file colors
    - Auto-adjusted column widths
    - Centered alignment for merged cells
    
    Args:
        results_df: Results DataFrame
        output_file: Output file path
        update_row_colors: Optional dict mapping row ID -> 6-char hex fill color (from update file)
    """
    if update_row_colors is None:
        update_row_colors = {}
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Alignment, Font
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.formatting.rule import FormulaRule
    
    logger.info("Creating formatted Excel output...")
    
    # Check if _youth_converted column exists
    has_youth_converted = '_youth_converted' in results_df.columns
    
    # Tag column: hide only when any product has "colosseo" in Product Tags
    hide_tag_column = bool(results_df.get('_has_colosseo_tag', pd.Series(dtype=bool)).fillna(False).any())
    
    # Columns to hide (not remove): ID and _from_update always; Tag when hide_tag_column; Colosseum-specific when applicable
    # Note: PNR and TIX NOM are NOT hidden - they appear after Total Units
    columns_always_hide = ['ID', '_from_update']
    if hide_tag_column:
        columns_always_hide.append('Tag')
    colosseum_only_hide = ['Sigilo', 'Codice', 'Ticket Group', 'Change By']
    has_colosseum_columns = any(col in results_df.columns for col in ['Codice', 'Sigilo', 'PNR'])
    
    # Remove columns that are completely empty (all NaN or empty strings)
    # But keep ID, _from_update, Tag, Notes (always show for all product tags), and Colosseum columns we want to hide (not remove)
    columns_to_never_remove = list(dict.fromkeys(
        columns_always_hide + ['Tag', 'Notes'] + (colosseum_only_hide + ['PNR', 'TIX NOM'] if has_colosseum_columns else [])
    ))
    columns_to_check = [col for col in results_df.columns 
                        if not col.startswith('_') and col not in columns_to_never_remove]
    empty_columns = []
    for col in columns_to_check:
        # Check if column is empty (all NaN, None, or empty strings)
        col_values = results_df[col].fillna('')
        if col_values.astype(str).str.strip().eq('').all():
            empty_columns.append(col)
    
    if empty_columns:
        logger.info(f"Removing empty columns from output: {empty_columns}")
        results_df = results_df.drop(columns=empty_columns)
    
    # Define desired column order
    # Order Reference after Full Name; Error, then Product (Ventrata), Product Code, then ID
    desired_column_order = [
        'Travel Date', 'Full Name', 'Order Reference', 'Unit Type', 'Total Units',
        'Tour Time', 'Language', 'Tour Type', 'Private Notes',
        'Change By', 'PNR', 'Ticket Group', 'Codice', 'Sigilo', 'TIX NOM',  # Colosseum columns
        'Error', 'Notes',
        'Product', 'Tag', 'Product Code','ID', 'Reseller', 'Public Notes'
    ]
    
    # Reorder columns: put known columns first in order, then any remaining columns
    existing_cols = list(results_df.columns)
    ordered_cols = [col for col in desired_column_order if col in existing_cols]
    # Add any remaining columns that aren't in the desired order (including internal _columns)
    remaining_cols = [col for col in existing_cols if col not in ordered_cols]
    final_column_order = ordered_cols + remaining_cols
    results_df = results_df[final_column_order]
    
    # Drop internal _has_colosseo_tag so it does not appear in Excel
    if '_has_colosseo_tag' in results_df.columns:
        results_df = results_df.drop(columns=['_has_colosseo_tag'])
    
    # Save basic Excel (including _youth_converted for now)
    results_df.to_excel(output_file, index=False)
    
    # Load workbook for formatting
    wb = load_workbook(output_file)
    ws = wb.active
    
    try:
        # Find column indices
        header_row = ws[1]
        col_indices = {}
        for idx, cell in enumerate(header_row, 1):
            if cell.value:
                col_indices[cell.value] = idx
        
        travel_date_col = col_indices.get('Travel Date')
        order_ref_col = col_indices.get('Order Reference')
        total_units_col = col_indices.get('Total Units')
        error_col = col_indices.get('Error')
        notes_col = col_indices.get('Notes')
        tour_time_col = col_indices.get('Tour Time')
        language_col = col_indices.get('Language')
        tour_type_col = col_indices.get('Tour Type')
        private_notes_col = col_indices.get('Private Notes')
        product_code_col = col_indices.get('Product Code')
        change_by_col = col_indices.get('Change By')
        codice_col = col_indices.get('Codice')
        sigilo_col = col_indices.get('Sigilo')
        reseller_col = col_indices.get('Reseller')
        
        # Format header row
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        for cell in header_row:
            if cell.value:  # Only format cells with content
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
        
        logger.info("Applied header formatting: blue background, white bold text")
        
        # Merge cells by Order Reference (group consecutive rows)
        # Also track booking ranges for zebra coloring
        booking_ranges = []  # List of (order_ref, start_row, end_row) tuples
        
        if order_ref_col:
            current_order_ref = None
            start_row = None
            
            for row_idx in range(2, ws.max_row + 2):  # +2 to include last group
                if row_idx <= ws.max_row:
                    cell_value = ws.cell(row=row_idx, column=order_ref_col).value
                else:
                    cell_value = None  # Force processing of last group
                
                if cell_value != current_order_ref:
                    # Process previous group
                    if current_order_ref is not None and start_row is not None:
                        end_row = row_idx - 1
                        
                        # Store booking range for zebra coloring
                        booking_ranges.append((current_order_ref, start_row, end_row))
                        
                        if end_row > start_row:  # Multiple rows in group
                            # Merge Order Reference
                            ws.merge_cells(start_row=start_row, start_column=order_ref_col,
                                         end_row=end_row, end_column=order_ref_col)
                            ws.cell(row=start_row, column=order_ref_col).alignment = Alignment(
                                horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=True)
                            
                            # Merge other shared columns
                            cols_to_merge = [
                                (travel_date_col, 'center'),
                                (total_units_col, 'center'),
                                (tour_time_col, 'center'),
                                (language_col, 'center'),
                                (tour_type_col, 'center'),
                                (private_notes_col, 'left'),
                                (product_code_col, 'left'),
                                (col_indices.get('Tag'), 'center'),
                                (change_by_col, 'center'),
                                (codice_col, 'center'),
                                (sigilo_col, 'center'),
                                (reseller_col, 'left'),
                                (error_col, 'left'),
                                (notes_col, 'left')
                            ]
                            
                            for col_idx, alignment in cols_to_merge:
                                if col_idx:
                                    ws.merge_cells(start_row=start_row, start_column=col_idx,
                                                 end_row=end_row, end_column=col_idx)
                                    ws.cell(row=start_row, column=col_idx).alignment = Alignment(
                                        horizontal=alignment, vertical='center', wrap_text=False, shrink_to_fit=True)
                    
                    # Start new group
                    if row_idx <= ws.max_row:
                        current_order_ref = cell_value
                        start_row = row_idx
        
        # Fixed row height and alignment so rows don't expand and long text shrinks to fit
        data_row_height = 22
        merged_top_left = set()  # (row, col) of top-left cell of each merged range
        for (_, start_row, end_row) in booking_ranges:
            if end_row > start_row:
                if order_ref_col:
                    merged_top_left.add((start_row, order_ref_col))
                for col_idx, _ in [
                    (travel_date_col, 'center'), (total_units_col, 'center'), (tour_time_col, 'center'),
                    (language_col, 'center'), (tour_type_col, 'center'), (private_notes_col, 'left'),
                    (product_code_col, 'left'), (col_indices.get('Tag'), 'center'), (change_by_col, 'center'),
                    (codice_col, 'center'), (sigilo_col, 'center'), (reseller_col, 'left'), (error_col, 'left'),
                    (notes_col, 'left')
                ]:
                    if col_idx:
                        merged_top_left.add((start_row, col_idx))
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = data_row_height
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) not in merged_top_left:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.alignment = Alignment(
                        horizontal='left', vertical='center', wrap_text=False, shrink_to_fit=True
                    )
        
        # Apply preserved row colors from update file first (before alternating colors)
        id_col_idx = col_indices.get('ID')
        rows_with_update_color = set()
        if update_row_colors and id_col_idx is not None and 'ID' in results_df.columns:
            for row_idx in range(2, ws.max_row + 1):
                df_idx = row_idx - 2
                if df_idx < 0 or df_idx >= len(results_df):
                    continue
                id_val = results_df.iloc[df_idx].get('ID')
                if pd.isna(id_val) or id_val == '':
                    continue
                id_key = str(id_val).strip()
                hex_color = update_row_colors.get(id_key)
                if hex_color:
                    rows_with_update_color.add(row_idx)
                    row_fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                    for col_idx in range(1, ws.max_column + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = row_fill
        
        # Apply alternating booking colors (gray for odd bookings, white for even)
        # Skip rows that have update-file colors so we don't overwrite them
        gray_fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        
        for booking_idx, (_, start_row, end_row) in enumerate(booking_ranges):
            # Alternate color: gray for odd bookings (0, 2, 4...), white for even (1, 3, 5...)
            if booking_idx % 2 == 0:  # Even index = odd booking number (1st, 3rd, 5th...)
                # Apply gray to all rows in this booking that don't have update-file color
                for row_idx in range(start_row, end_row + 1):
                    if row_idx in rows_with_update_color:
                        continue
                    for col_idx in range(1, len(results_df.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        try:
                            no_fill = (getattr(cell.fill, 'fill_type', None) is None or
                                       (getattr(cell.fill.start_color, 'index', None) == '00000000'))
                        except Exception:
                            no_fill = True
                        if no_fill:
                            cell.fill = gray_fill
        
        # Highlight rows where Youth was converted to Adult (non-EU) in faded yellow
        # Skip rows that have preserved update-file colors so we don't overwrite them
        if has_youth_converted:
            youth_converted_col = col_indices.get('_youth_converted')
            if youth_converted_col:
                faded_yellow_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")
                
                for row_idx in range(2, ws.max_row + 1):
                    if row_idx in rows_with_update_color:
                        continue
                    converted_cell = ws.cell(row=row_idx, column=youth_converted_col)
                    if converted_cell.value in [True, 'True', 'TRUE', 1]:
                        # Highlight entire row in faded yellow
                        for col_idx in range(1, len(results_df.columns) + 1):
                            # Skip the _youth_converted column itself
                            if col_idx != youth_converted_col:
                                ws.cell(row=row_idx, column=col_idx).fill = faded_yellow_fill
        
        # Highlight rows with errors in bright yellow (overrides faded yellow)
        # Skip rows that have preserved update-file colors so we don't overwrite them
        if error_col:
            yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            for row_idx in range(2, ws.max_row + 1):
                if row_idx in rows_with_update_color:
                    continue
                error_cell = ws.cell(row=row_idx, column=error_col)
                if error_cell.value and str(error_cell.value).strip():
                    # Highlight entire row
                    for col_idx in range(1, len(results_df.columns) + 1):
                        ws.cell(row=row_idx, column=col_idx).fill = yellow_fill
        
        # Highlight Unit Type cells for Child/Infant in light blue (skip rows with preserved colors)
        unit_type_col = col_indices.get('Unit Type')
        if unit_type_col:
            light_blue_fill = PatternFill(start_color="89CFF0", end_color="89CFF0", fill_type="solid")
            yellow_fill_youth = PatternFill(start_color="FFBF00", end_color="FFBF00", fill_type="solid")
            
            for row_idx in range(2, ws.max_row + 1):
                if row_idx in rows_with_update_color:
                    continue
                unit_type_cell = ws.cell(row=row_idx, column=unit_type_col)
                unit_value = str(unit_type_cell.value).strip() if unit_type_cell.value else ''
                
                if unit_value in ['Child', 'Infant']:
                    # Highlight Child/Infant in light blue
                    unit_type_cell.fill = light_blue_fill
                elif unit_value == 'Youth':
                    # Highlight Youth in yellow (same as error highlighting)
                    unit_type_cell.fill = yellow_fill_youth
        
        # Apply Tag dropdowns and conditional coloring per booking
        tag_col = col_indices.get('Tag')
        has_tag_options_column = '_tag_options' in results_df.columns
        if tag_col and has_tag_options_column and booking_ranges:
            tag_column_letter = ws.cell(row=1, column=tag_col).column_letter
            
            for order_ref, start_row, _ in booking_ranges:
                df_idx = start_row - 2  # account for header row
                if df_idx < 0 or df_idx >= len(results_df):
                    continue
                
                options = results_df.iloc[df_idx].get('_tag_options', [])
                if not isinstance(options, list) or not options:
                    continue
                
                labels = []
                color_map = {}
                for option in options:
                    label = option.get('label')
                    if not label:
                        continue
                    if label not in labels:
                        labels.append(label)
                    color = option.get('color')
                    if color:
                        color_map[label] = color
                
                if not labels:
                    continue
                
                sanitized = [label.replace('"', '""') for label in labels]
                joined = ",".join(sanitized)
                dv = DataValidation(type="list", formula1=f'"{joined}"', allow_blank=True)
                ws.add_data_validation(dv)
                
                cell_address = f"{tag_column_letter}{start_row}"
                tag_cell = ws[cell_address]
                dv.add(tag_cell)
                
                # Apply conditional formatting per label on the merged tag cell
                for label, color in color_map.items():
                    if not color:
                        continue
                    fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    formula = f'{cell_address}="{label}"'
                    ws.conditional_formatting.add(cell_address, FormulaRule(formula=[formula], fill=fill, stopIfTrue=False))
        
        # Remove _youth_converted column if it exists (internal flag, not for user)
        if has_youth_converted:
            youth_converted_col = col_indices.get('_youth_converted')
            if youth_converted_col:
                ws.delete_cols(youth_converted_col)
                logger.info("Removed internal _youth_converted column from Excel output")
        
        # Remove internal _tag_options column if present
        if has_tag_options_column:
            tag_options_col = None
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == '_tag_options':
                    tag_options_col = idx
                    break
            if tag_options_col:
                ws.delete_cols(tag_options_col)
                logger.info("Removed internal _tag_options column from Excel output")
        
        # Auto-adjust column widths (after removing _youth_converted)
        for col_idx, col in enumerate(results_df.columns, 1):
            # Skip internal columns
            if col in ['_youth_converted', '_tag_options']:
                continue
            max_length = max(
                results_df[col].astype(str).apply(len).max(),
                len(str(col))
            ) + 2
            
            # Set minimum width of 20 for PNR, Codice, Sigilo columns
            if col in ['PNR', 'Codice', 'Sigilo','Public Notes']:
                max_length = max(max_length, 20)
            
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            ws.column_dimensions[col_letter].width = min(max_length, 50)
        
        # Header row height (data row heights set earlier with fixed height + shrink-to-fit)
        ws.row_dimensions[1].height = 20
        
        # Hide columns: always ID and _from_update; Tag when any product has colosseo; Colosseum-specific when applicable
        columns_to_hide = list(columns_always_hide)
        if has_colosseum_columns:
            columns_to_hide = columns_to_hide + colosseum_only_hide
        hidden_cols = []
        for col_name in columns_to_hide:
            if col_name in col_indices:
                col_idx = col_indices[col_name]
                col_letter = ws.cell(row=1, column=col_idx).column_letter
                ws.column_dimensions[col_letter].hidden = True
                hidden_cols.append(col_name)
        if hidden_cols:
            logger.info(f"Hidden columns in Excel: {hidden_cols}")
        
        # Freeze header row (keep it visible when scrolling)
        ws.freeze_panes = 'A2'  # Freeze first row (header)
        logger.info("Applied freeze panes: header row will stay visible when scrolling")
        
        # Save formatted workbook
        wb.save(output_file)
        logger.info("Applied Excel formatting: merged cells, colors, auto-widths, freeze panes")
        
    except Exception as e:
        logger.warning(f"Could not apply Excel formatting: {e}")
        logger.info("Basic Excel file saved without formatting")
    finally:
        # Remove internal helper column from the DataFrame so callers don't see it
        if '_tag_options' in results_df.columns:
            results_df.drop(columns=['_tag_options'], inplace=True)


def main():
    """Main entry point for name extraction."""
    logger.info("=" * 80)
    logger.info("Names Generation System - Starting")
    logger.info("=" * 80)
    
    # Example usage - can be modified for CLI arguments or GUI later
    # For now, using hardcoded paths for testing
    
    # TODO: Replace with actual file paths or CLI arguments
    ventrata_file = "/Users/navidabasi/Downloads/tickets.xlsx"
    monday_file = "path/to/monday.xlsx"  # Optional
    update_file = "path/to/update.xlsx"  # Optional - previously extracted data
    
    try:
        # Step 1: Load data
        logger.info("\n" + "=" * 80)
        logger.info("STEP 1: Loading Data Files")
        logger.info("=" * 80)
        
        if not os.path.exists(ventrata_file):
            logger.error(f"Ventrata file not found: {ventrata_file}")
            logger.info("\nPlease update the file paths in main.py")
            logger.info("Example:")
            logger.info('  ventrata_file = "/Users/username/path/to/ventrata.xlsx"')
            logger.info('  monday_file = "/Users/username/path/to/monday.xlsx"  # Optional')
            return
        
        ventrata_df = load_ventrata(ventrata_file)
        
        # Load Monday file if it exists
        monday_df = None
        if monday_file and os.path.exists(monday_file):
            logger.info("Monday file provided - will merge with Ventrata data")
            monday_df = load_monday(monday_file)
        else:
            logger.info("No Monday file - processing Ventrata only")
        
        # Load Update file if it exists
        update_df = None
        update_row_colors = {}
        if update_file and os.path.exists(update_file):
            logger.info("Update file provided - will reuse previously extracted names")
            update_df, update_row_colors = load_update_file(update_file)
        else:
            logger.info("No Update file - extracting all names from scratch")
        
        # Step 2: Merge data if Monday provided
        if monday_df is not None:
            logger.info("\n" + "=" * 80)
            logger.info("STEP 2: Merging Ventrata and Monday Data")
            logger.info("=" * 80)
            merged_df = merge_data(ventrata_df, monday_df)
        else:
            merged_df = ventrata_df
        
        # Step 3: Process names
        logger.info("\n" + "=" * 80)
        logger.info("STEP 3: Extracting Names and Validating Data")
        logger.info("=" * 80)
        
        processor = NameExtractionProcessor(merged_df, monday_df, update_df)
        results_df = processor.process()
        
        # Step 4: Display results
        logger.info("\n" + "=" * 80)
        logger.info("STEP 4: Processing Complete")
        logger.info("=" * 80)
        
        logger.info(f"\nTotal entries processed: {len(results_df)}")
        
        if not results_df.empty:
            # Count entries with errors
            errors_count = results_df['Error'].ne('').sum()
            logger.info(f"Entries with errors: {errors_count}")
            logger.info(f"Entries without errors: {len(results_df) - errors_count}")
            
            # Show sample of results
            logger.info("\nSample results (first 5 entries):")
            logger.info("-" * 80)
            sample_cols = ['Full Name', 'Order Reference', 'Unit Type', 'Reseller', 'Error']
            available_cols = [col for col in sample_cols if col in results_df.columns]
            print(results_df[available_cols].head().to_string(index=False))
            
            # Show error statistics if any
            if errors_count > 0:
                logger.info("\nError breakdown:")
                logger.info("-" * 80)
                error_df = results_df[results_df['Error'] != '']
                # Count unique error types
                error_types = {}
                for error in error_df['Error']:
                    for err in str(error).split(' | '):
                        err = err.strip()
                        if err:
                            error_types[err] = error_types.get(err, 0) + 1
                
                for error_type, count in sorted(error_types.items(), key=lambda x: -x[1]):
                    logger.info(f"  {error_type}: {count}")
            
            # Save to Excel with formatting (auto-increment filename if exists)
            base_output_file = "names_output.xlsx"
            output_file = get_next_available_filename(base_output_file)
            
            if output_file != base_output_file:
                logger.info(f"\n{base_output_file} already exists, using: {output_file}")
            else:
                logger.info(f"\nSaving results to: {output_file}")
            
            save_results_to_excel(results_df, output_file, update_row_colors=update_row_colors)
            logger.info(f"Results saved successfully!")
            
            # Try to open the Excel file automatically
            try:
                import platform
                import subprocess
                
                if platform.system() == 'Darwin':  # macOS
                    subprocess.run(['open', output_file])
                elif platform.system() == 'Windows':
                    os.startfile(output_file)
                else:  # Linux
                    subprocess.run(['xdg-open', output_file])
                
                logger.info(f"Opened {output_file} in Excel")
            except Exception as e:
                logger.info(f"Could not auto-open file (you can open it manually): {e}")
        else:
            # Empty results - still save file
            logger.warning("No data was extracted!")
            base_output_file = "names_output.xlsx"
            output_file = get_next_available_filename(base_output_file)
            results_df.to_excel(output_file, index=False)
            logger.info(f"Empty results file saved to: {output_file}")
        
        logger.info("\n" + "=" * 80)
        logger.info(f"📊 EXCEL FILE: {os.path.abspath(output_file)}")
        logger.info("Process completed successfully!")
        logger.info("=" * 80)
        
        return results_df
        
    except FileNotFoundError as e:
        logger.error(f"File not found: {e}")
        logger.info("\nPlease check your file paths and try again.")
        sys.exit(1)
    except Exception as e:
        logger.error(f"Error during processing: {e}")
        import traceback
        logger.error(traceback.format_exc())
        sys.exit(1)


if __name__ == "__main__":
    main()
